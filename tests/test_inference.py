import unittest
import opennre
import json
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment


wb = Workbook()
wb.create_sheet('分类准确率，召回率，f1值')
ws = wb['Sheet']
ws1 = wb['分类准确率，召回率，f1值']
data = [[0, 0, 0, 0] for _ in range(9)]
red_fill = PatternFill("solid", fgColor="CCCCFF")  # 单元格填充颜色
line_count = 414


class TestInference(unittest.TestCase):

    def test_wiki80_cnn_softmax(self):
        # TODO 测试时这里 get_model 中的路径有很多坑，要注意！！！！，每次运行前，记得把模型拷到 .opennre 文件夹中的对应位置。
        model_name = 'dde_bert-base-uncased_entity'
        shutil.copy('../model/ckpt/dde_bert-base-uncased_entity.pth.tar', '../.opennre/pretrain/nre/dde_bert-base-uncased_entity.pth.tar')
        model = opennre.get_model(model_name)  # 使用 dde cnn 来训练数据
        sens = []
        with open('../data/dde/dde_val.txt', encoding='utf-8') as f:
            for i in range(line_count):
                print(i)
                foo = eval(f.readline())
                line = json.dumps(foo)
                sens.append(json.loads(line))
        row = 1
        col = 1
        ws.cell(row, col).value = '实体1'
        ws.cell(row, col + 1).value = '实体2'
        ws.cell(row, col + 2).value = '实际关系'
        ws.cell(row, col + 3).value = '预测关系'
        ws.cell(row, col + 4).value = '概率'
        ws.cell(row, col + 5).value = '原始句子'
        for sen in sens:
            row += 1
            result = model.infer(sen)
            print('识别句子：', ' '.join(sen['token']))
            print('实际关系三元组：', sen['h']['name'] + ' / ' + sen['t']['name'] + ' / ' + sen['relation'])
            print('预测关系和概率：', result)
            ws.cell(row, col).value = sen['h']['name']
            ws.cell(row, col+1).value = sen['t']['name']
            ws.cell(row, col+2).value = sen['relation']
            ws.cell(row, col+3).value = result[0]
            ws.cell(row, col+4).value = result[1]
            ws.cell(row, col+5).value = ' '.join(sen['token'])
            if sen['relation'] != result[0]:
                ws.cell(row, col+3).fill = red_fill
            word_tags = [0, sen['relation'], result[0]]
            if 'Cause-Effect' in word_tags[1]:
                data[0][0] += 1
                if word_tags[1] == word_tags[2]:
                    data[0][1] += 1
                elif word_tags[2] != 'O':
                    data[0][2] += 1
                else:
                    data[0][3] += 1
            if 'Instrument-Agency' in word_tags[1]:
                data[1][0] += 1
                if word_tags[1] == word_tags[2]:
                    data[1][1] += 1
                elif word_tags[2] != 'O':
                    data[1][2] += 1
                else:
                    data[1][3] += 1
            if 'Component-Whole' in word_tags[1]:
                data[2][0] += 1
                if word_tags[1] == word_tags[2]:
                    data[2][1] += 1
                elif word_tags[2] != 'O':
                    data[2][2] += 1
                else:
                    data[2][3] += 1
            if 'Entity-Destination' in word_tags[1]:
                data[3][0] += 1
                if word_tags[1] == word_tags[2]:
                    data[3][1] += 1
                elif word_tags[2] != 'O':
                    data[3][2] += 1
                else:
                    data[3][3] += 1
            if 'Member-Collection' in word_tags[1]:
                data[4][0] += 1
                if word_tags[1] == word_tags[2]:
                    data[4][1] += 1
                elif word_tags[2] != 'O':
                    data[4][2] += 1
                else:
                    data[4][3] += 1
            if 'Entity-Origin' in word_tags[1]:
                data[5][0] += 1
                if word_tags[1] == word_tags[2]:
                    data[5][1] += 1
                elif word_tags[2] != 'O':
                    data[5][2] += 1
                else:
                    data[5][3] += 1
            if 'Product-Producer' in word_tags[1]:
                data[6][0] += 1
                if word_tags[1] == word_tags[2]:
                    data[6][1] += 1
                elif word_tags[2] != 'O':
                    data[6][2] += 1
                else:
                    data[6][3] += 1
            if 'Content-Container' in word_tags[1]:
                data[7][0] += 1
                if word_tags[1] == word_tags[2]:
                    data[7][1] += 1
                elif word_tags[2] != 'O':
                    data[7][2] += 1
                else:
                    data[7][3] += 1
            if 'Message-Topic' in word_tags[1]:
                data[8][0] += 1
                if word_tags[1] == word_tags[2]:
                    data[8][1] += 1
                elif word_tags[2] != 'O':
                    data[8][2] += 1
                else:
                    data[8][3] += 1
        ws1.cell(1, 1).value = '类别'
        ws1.cell(1, 2).value = '准确率'
        ws1.cell(1, 3).value = '召回率'
        ws1.cell(1, 4).value = 'f1 值'
        ws1.cell(2, 1).value = 'Cause-Effect'
        ws1.cell(3, 1).value = 'Instrument-Agency'
        ws1.cell(4, 1).value = 'Component-Whole'
        ws1.cell(5, 1).value = 'Entity-Destination'
        ws1.cell(6, 1).value = 'Member-Collection'
        ws1.cell(7, 1).value = 'Entity-Origin'
        ws1.cell(8, 1).value = 'Product-Producer'
        ws1.cell(9, 1).value = 'Content-Container'
        ws1.cell(10, 1).value = 'Message-Topic'

        for i in range(9):
            if data[i][1] == 0:
                ws1.cell(i + 2, 2).value = 0
                ws1.cell(i + 2, 3).value = 0
                ws1.cell(i + 2, 4).value = 0
            else:
                ws1.cell(i + 2, 2).value = data[i][1] / (data[i][1] + data[i][2])
                p = data[i][1] / (data[i][1] + data[i][2])
                ws1.cell(i + 2, 3).value = data[i][1] / (data[i][1] + data[i][3])
                r = data[i][1] / (data[i][1] + data[i][3])
                ws1.cell(i + 2, 4).value = 2 * p * r / (p + r)

        wb.save('../tools/out_data/res_excel.xlsx')


if __name__ == '__main__':
    unittest.main()
