from openpyxl import load_workbook
import json

# 用来把地学院的excel数据转换成json

wb = load_workbook('./raw_data/raw_dde.xlsx')
ws = wb['原始筛选']

train = []  # 训练集
test = []  # 测试集
words_vec2id = dict()  # 存放所有词向量的id
words2id = dict()  # 存放出现的词向量的id
all_relation = set()  # 存放所有出现过的关系
begin, end = 2, 319  # 数据起始和结束的行号 2 --> 319
train_f = open('../data/dde/dde_train.txt', 'w', encoding='utf-8')
test_f = open('../data/dde/dde_val.txt', 'w',  encoding='utf-8')

# 遍历句子生成 json
divider = 1
for i in range(begin, end):
    line_number = ws['A' + str(i)].value
    instance1 = ws['B' + str(i)].value.replace(',', ' ,').replace('.', ' .').strip()  # 去掉多余的空格。手动将,和.分开
    instance1_type = ws['C' + str(i)].value
    instance2 = ws['D' + str(i)].value.replace(',', ' ,').replace('.', ' .').strip()
    instance2_type = ws['E' + str(i)].value
    relation = ws['F' + str(i)].value
    sentence = ws['H' + str(i)].value.lower().replace(',', ' ,').replace('.', ' .').strip()  # 原始英文字符串，这里将所有的大写字母转换成小写，保证能够找到词向量
    trigger_word = ws['G' + str(i)].value
    seg_sen = sentence.split(' ')  # 分词后将所有词包装成一个列表
    ori_sen = sentence  # 无空格的句子 (中文中是如此，但是在英文里就没有必要了，所以这里还是保持是原来的句子)
    print(seg_sen)
    try:
        instance1_pos = []
        instance1_seg = instance1.lower().split(' ')
        instance1_pos.append(seg_sen.index(instance1_seg[0]))
        instance1_pos.append(seg_sen.index(instance1_seg[-1]) + 1)  # 区间是左闭右开，所以要加一
        # print(instance1_pos)
        instance2_pos = []
        instance2_seg = instance2.lower().split(' ')
        instance2_pos.append(seg_sen.index(instance2_seg[0]))
        instance2_pos.append(seg_sen.index(instance2_seg[-1]) + 1)  # 区间是左闭右开，所以要加一
        # print(instance2_pos)
    except Exception as e:
        print('错误！错误行号：', line_number, e)
        continue

    row = {
        'token': seg_sen,
        'h': {
            'name': instance1.lower(),
            'pos': instance1_pos,  # 这个单词组在，分割句子中的位置
            'id': '0'  # 不知道这个参数有什么用
        },
        't': {
            'name': instance2.lower(),
            'pos': instance2_pos,
            'id': '0'
        },
        'relation': relation
    }

    all_relation.add(relation)  # 统计出现的关系种类

    divider += 1
    if divider % 5 != 0:  # 目前有效条目 104， 训练集83，测试集21 5:1 均匀分配
        train.append(row)
        train_f.write(str(row))
        train_f.write('\n')
    else:
        test.append(row)
        test_f.write(str(row))
        test_f.write('\n')

train_f.close()
test_f.close()
print(all_relation)  # 测试用print
print(len(train), len(test))

# with open('./raw_data/train_dde.json', 'w') as f:  # 写入json文件
#     json.dump(train, f)
# with open('./raw_data/test_dde.json', 'w') as f:
#     json.dump(test, f)
